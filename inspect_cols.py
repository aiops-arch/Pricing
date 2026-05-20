import openpyxl

def show_all_cols(path, label):
    print("\n" + "-"*80)
    print("FILE: " + label)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        print("Total cols: " + str(ws.max_column) + "  Total rows: " + str(ws.max_row))
        # Row 1 = headers
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row)
        print("ALL COLUMN HEADERS:")
        for i, h in enumerate(headers):
            print("  col" + str(i+1).zfill(3) + ": " + str(h))
        # Show a couple data rows
        print("\nSAMPLE DATA ROW 2:")
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            for i, val in enumerate(row):
                if val is not None:
                    print("  col" + str(i+1).zfill(3) + " [" + str(headers[i]) + "]: " + str(val))
        wb.close()
    except Exception as e:
        print("ERROR: " + str(e))

show_all_cols(
    r"E:\Pricing\Monthly\MAY-2026\0.30UP POSITION REPORT - 15-05-2026.xlsx",
    "POSITION REPORT 15-05-2026"
)
