import openpyxl

def inspect(path, label, max_rows=12):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print()
    print("-" * 90)
    print("FILE: " + label)
    print("Sheets: " + str(wb.sheetnames))
    print("Dims: rows=" + str(ws.max_row) + "  cols=" + str(ws.max_column))
    print("First " + str(max_rows) + " rows:")
    for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        print("  row" + str(i+1).zfill(2) + ": " + str(list(row)))

inspect(r"E:\Pricing\FANCY BASE REPORT 23-03-2026.xlsx", "FANCY BASE REPORT 23-03-2026")
inspect(r"E:\Pricing\ASSCHER-HEART BASE REPORT 23-03-2026.xlsx", "ASSCHER-HEART BASE REPORT 23-03-2026")
inspect(r"E:\Pricing\Monthly\MAY-2026\0.30UP POSITION REPORT - 15-05-2026.xlsx", "0.30UP POSITION REPORT 15-05-2026 (latest May)")
inspect(r"E:\Pricing\Monthly\JAN-2026\0.30UP POSITION REPORT 01-01-2026.xlsx", "0.30UP POSITION REPORT 01-01-2026 (oldest Jan)")
