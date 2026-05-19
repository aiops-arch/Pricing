"""
loader.py
---------
Loads the Fancy Base Report and Asscher-Heart Base Report Excel files.

The Base Reports have a multi-row header structure (rows 1-4 are header rows,
data starts at row 5). This module reads the raw workbook, reconstructs a clean
single-row header by collapsing the merged header rows, and returns a tidy
DataFrame ready for normalization.
"""

import logging
import os
import re
from pathlib import Path
from typing import Union

import numpy as np
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# Number of header rows in the base report Excel files
HEADER_ROWS = 4


def _collapse_header_cell(values: list) -> str:
    """
    Collapse multiple header row values into a single column name by joining
    non-null, non-empty values with an underscore separator.

    Parameters
    ----------
    values : list
        List of cell values from header rows for a single column.

    Returns
    -------
    str
        Collapsed header string, e.g. "Comp_1_Disc%".
    """
    parts = []
    for v in values:
        if v is not None and str(v).strip() not in ("", "None"):
            parts.append(str(v).strip())
    return "_".join(parts) if parts else "unknown"


def _read_raw_headers(ws, header_rows: int = HEADER_ROWS) -> list:
    """
    Read the raw header rows from an openpyxl worksheet, resolving merged cells
    by propagating the top-left value of each merged range.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The active worksheet object.
    header_rows : int
        Number of rows to treat as header (default 4).

    Returns
    -------
    list of list
        A list of ``header_rows`` lists, each containing the cell values for
        every column in that header row.
    """
    max_col = ws.max_column

    # Build a dict of merged-cell fill values: (row, col) -> value
    merged_fill: dict = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_fill[(row, col)] = top_left_value

    raw_headers = []
    for row_idx in range(1, header_rows + 1):
        row_values = []
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in merged_fill:
                row_values.append(merged_fill[(row_idx, col_idx)])
            else:
                row_values.append(ws.cell(row_idx, col_idx).value)
        raw_headers.append(row_values)

    return raw_headers


def _build_column_names(raw_headers: list) -> list:
    """
    Build a flat list of column names by collapsing the multi-row header for
    each column position.

    Duplicate column names are de-duplicated by appending a numeric suffix.

    Parameters
    ----------
    raw_headers : list of list
        Output from ``_read_raw_headers``.

    Returns
    -------
    list of str
        One column name per column in the spreadsheet.
    """
    if not raw_headers:
        return []

    num_cols = len(raw_headers[0])
    column_names = []
    seen: dict = {}

    for col_idx in range(num_cols):
        cell_values = [row[col_idx] for row in raw_headers]
        name = _collapse_header_cell(cell_values)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        column_names.append(name)

    return column_names


def _detect_report_type(path: Union[str, Path]) -> str:
    """
    Detect whether the file is a FANCY or ASSCHER-HEART base report by
    inspecting the filename.

    Parameters
    ----------
    path : str or Path
        Full path to the Excel file.

    Returns
    -------
    str
        ``"FANCY"`` or ``"ASSCHER"`` or ``"UNKNOWN"``.
    """
    filename = Path(path).name.upper()
    if "ASSCHER" in filename or "HEART" in filename:
        return "ASSCHER"
    if "FANCY" in filename:
        return "FANCY"
    return "UNKNOWN"


def _parse_report_date(path: Union[str, Path]) -> str:
    """
    Attempt to extract a date string from the filename.  Looks for patterns
    like ``2024-05-13``, ``13-05-2024``, ``05132024``, or ``13052024``.

    Parameters
    ----------
    path : str or Path
        Full path to the Excel file.

    Returns
    -------
    str
        ISO-format date string ``"YYYY-MM-DD"`` if found, otherwise ``""`` .
    """
    filename = Path(path).stem
    # Try YYYY-MM-DD or YYYY_MM_DD
    m = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})", filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # Try DD-MM-YYYY or DD_MM_YYYY
    m = re.search(r"(\d{2})[-_](\d{2})[-_](\d{4})", filename)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    # Try DDMMYYYY or MMDDYYYY (8 consecutive digits)
    m = re.search(r"(\d{8})", filename)
    if m:
        s = m.group(1)
        # Assume DDMMYYYY
        try:
            day, month, year = int(s[:2]), int(s[2:4]), int(s[4:])
            if 1 <= month <= 12 and 1 <= day <= 31:
                return f"{year:04d}-{month:02d}-{day:02d}"
        except ValueError:
            pass
    logger.warning("Could not parse report_date from filename: %s", filename)
    return ""


def load_base_report(path: Union[str, Path]) -> pd.DataFrame:
    """
    Load a Fancy or Asscher-Heart Base Report Excel file.

    The file is expected to have a 4-row merged header followed by data rows
    starting at row 5.  This function:

    1. Opens the workbook with openpyxl to resolve merged header cells.
    2. Builds clean column names by collapsing the 4-row header.
    3. Reads all data rows (row 5 onward) into a DataFrame.
    4. Strips whitespace from all string values.
    5. Adds metadata columns ``source_file`` and ``report_date``.

    Parameters
    ----------
    path : str or pathlib.Path
        Full path to the ``.xlsx`` base report file.

    Returns
    -------
    pd.DataFrame
        Cleaned DataFrame with standardised column names.  Contains two extra
        columns: ``source_file`` (filename) and ``report_date`` (ISO date str).

    Raises
    ------
    FileNotFoundError
        If the file does not exist at ``path``.
    ValueError
        If the file appears to have no data rows.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Base report not found: {path}")

    logger.info("Loading base report: %s", path.name)

    # ------------------------------------------------------------------ #
    # Step 1: Use openpyxl to handle merged headers                        #
    # ------------------------------------------------------------------ #
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    raw_headers = _read_raw_headers(ws, HEADER_ROWS)
    column_names = _build_column_names(raw_headers)
    logger.debug("Detected %d columns in header", len(column_names))

    # ------------------------------------------------------------------ #
    # Step 2: Read data rows (row 5 onward) from openpyxl                  #
    # ------------------------------------------------------------------ #
    data_rows = []
    for row in ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue
        data_rows.append(list(row))

    if not data_rows:
        raise ValueError(f"No data rows found in {path.name} (expected data from row {HEADER_ROWS + 1})")

    # ------------------------------------------------------------------ #
    # Step 3: Build DataFrame                                              #
    # ------------------------------------------------------------------ #
    # Pad or trim column names to match actual column count
    num_cols_in_data = len(data_rows[0]) if data_rows else len(column_names)
    if len(column_names) < num_cols_in_data:
        for i in range(len(column_names), num_cols_in_data):
            column_names.append(f"col_{i}")
    elif len(column_names) > num_cols_in_data:
        column_names = column_names[:num_cols_in_data]

    df = pd.DataFrame(data_rows, columns=column_names)

    # ------------------------------------------------------------------ #
    # Step 4: Strip whitespace from all string values                      #
    # ------------------------------------------------------------------ #
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)

    # ------------------------------------------------------------------ #
    # Step 5: Add metadata columns                                         #
    # ------------------------------------------------------------------ #
    df["source_file"] = path.name
    df["report_date"] = _parse_report_date(path)
    report_type = _detect_report_type(path)
    df["report_type"] = report_type

    logger.info(
        "Loaded %d rows from %s (type=%s, report_date=%s)",
        len(df),
        path.name,
        report_type,
        df["report_date"].iloc[0] if len(df) > 0 else "N/A",
    )
    return df
