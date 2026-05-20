"""Inspect xlsx files using only openpyxl (no numpy/pandas to avoid OpenBLAS crash)."""
import openpyxl

def inspect_sheet(path, sheet_index=0, max_rows=12):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_index]
    print(f"\n{'='*80}")
    print(f"FILE: {path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Active sheet: {ws.title}  |  dims: {ws.dimensions}  |  rows={ws.max_row} cols={ws.max_column}")
    print(f"--- First {max_rows} rows ---")
    for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        print(f"  row{i+1:02d}: {row}")

inspect_sheet(r"FANCY BASE REPORT 23-03-2026.xlsx", max_rows=15)
inspect_sheet(r"ASSCHER-HEART BASE REPORT 23-03-2026.xlsx", max_rows=15)

# DANY - check both sheets
wb3 = openpyxl.load_workbook(r"DANY ORDER LIST.xlsx", data_only=True)
print(f"\n{'='*80}")
print(f"FILE: DANY ORDER LIST.xlsx  |  Sheets: {wb3.sheetnames}")
for sname in wb3.sheetnames:
    ws = wb3[sname]
    print(f"\n  Sheet: '{sname}'  rows={ws.max_row} cols={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        print(f"    row{i+1:02d}: {row}")
