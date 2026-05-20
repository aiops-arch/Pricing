import openpyxl

# Inspect just the Fancy Base Report - headers + first data rows
path = r"E:\Pricing\FANCY BASE REPORT 23-03-2026.xlsx"
print("Opening Fancy Base Report (read_only mode)...")
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb.active
print("Sheets: " + str(wb.sheetnames))
print("Max rows: " + str(ws.max_row) + "  Max cols: " + str(ws.max_column))

print("\n--- HEADER ROWS 1-4 (first 30 cols only) ---")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True)):
    print("row" + str(i+1) + ": " + str(list(row)[:30]))

print("\n--- DATA ROWS 5-14 (first 30 cols) ---")
for i, row in enumerate(ws.iter_rows(min_row=5, max_row=14, values_only=True)):
    print("row" + str(i+5) + ": " + str(list(row)[:30]))

wb.close()
print("\nDone.")
