import openpyxl

path = r"E:\Pricing\Monthly\MAY-2026\0.30UP POSITION REPORT - 15-05-2026.xlsx"
print("Inspecting Position Report for ROUND 0.30-2.00ct stones...")
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb.active

# Get headers from row 1
headers = []
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    headers = list(row)

# Find col indices for key fields
def col_idx(name):
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None

shape_col = col_idx('Shape')
cts_col = col_idx('cts')
color_col = col_idx('Color')
clarity_col = col_idx('Clarity')
cut_col = col_idx('Cut')
polish_col = col_idx('Polish')
sym_col = col_idx('Symmetry')
fluor_col = col_idx('Fluorescence')
loc_col = col_idx('Location')
lab_col = col_idx('Lab')
stone_id_col = col_idx('Stone Id')
inward_col = col_idx('Inward date')
aging_col = col_idx('AgingDays')
mail_back_col = col_idx('Mail Back')
final_disc_col = col_idx('Final disc')
rapnet_disc_col = col_idx('Rapnet disc +')
real_rapnet_col = col_idx('REAL RAPNET')
tariff_rapnet_col = col_idx('Tariff Rapnet +')
base_pd_disc_col = col_idx('Base pd disc')
rapnet_pos_ind_col = col_idx('Rapnet Pos IND')
rapnet_pos_world_col = col_idx('Rapnet Pos')
rapnet_pos_usa_col = col_idx('Rapnet Pos USA')
rapnet_pcs_ind_col = col_idx('Rapnet Pcs Pos IND')
first_comp_col = col_idx('1st')  # first competitor position (World)
limit_col = col_idx('LIMIT 1')
limit_remark_col = col_idx('LIMIT REMARK')
status_col = col_idx('Stone status')
inv_remark_col = col_idx('RAPNET_STATUS')
mixstart_col = col_idx('Mix StartDiscountPercent')
avg_disc_col = col_idx('AvgDiscountPercent')
rapsize_col = col_idx('RAPNET_SIZE')
size_col = col_idx('Psize')
rapnet_aging_col = col_idx('RAPNET AGING')
days_rapnet_col = headers.index('Days On Rapnet in Last (639 Days)') if 'Days On Rapnet in Last (639 Days)' in headers else None

print("Key column positions:")
print("  Shape=" + str(shape_col) + "  cts=" + str(cts_col) + "  Color=" + str(color_col) + "  Clarity=" + str(clarity_col))
print("  Cut=" + str(cut_col) + "  Polish=" + str(polish_col) + "  Symmetry=" + str(sym_col) + "  Fluor=" + str(fluor_col))
print("  Location=" + str(loc_col) + "  Lab=" + str(lab_col) + "  AgingDays=" + str(aging_col))
print("  Mail Back=" + str(mail_back_col) + "  Final disc=" + str(final_disc_col))
print("  Rapnet disc+=" + str(rapnet_disc_col) + "  REAL RAPNET=" + str(real_rapnet_col))
print("  Tariff Rapnet+=" + str(tariff_rapnet_col) + "  Base pd disc=" + str(base_pd_disc_col))
print("  LIMIT 1=" + str(limit_col) + "  LIMIT REMARK=" + str(limit_remark_col))
print("  RapPos IND=" + str(rapnet_pos_ind_col) + "  RapPos World=" + str(rapnet_pos_world_col) + "  RapPos USA=" + str(rapnet_pos_usa_col))
print("  1st competitor=" + str(first_comp_col) + "  AvgDisc=" + str(avg_disc_col) + "  Psize=" + str(size_col))

# Collect ROUND 0.30-2.00 rows
round_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if shape_col is not None and row[shape_col] == 'ROUND':
        cts = row[cts_col] if cts_col is not None else None
        if cts is not None:
            try:
                cts_f = float(cts)
                if 0.30 <= cts_f <= 2.00:
                    round_rows.append(row)
            except:
                pass

print("\nROUND 0.30-2.00ct stones in this file: " + str(len(round_rows)))

# Show 5 samples with key columns
print("\nSAMPLE ROUND STONES (key columns):")
for r in round_rows[:8]:
    print()
    print("  Stone: " + str(r[stone_id_col] if stone_id_col else '?') +
          "  cts=" + str(r[cts_col] if cts_col else '?') +
          "  RapSize=" + str(r[rapsize_col] if rapsize_col else '?') +
          "  Psize=" + str(r[size_col] if size_col else '?'))
    print("  Color=" + str(r[color_col] if color_col else '?') +
          "  Clarity=" + str(r[clarity_col] if clarity_col else '?') +
          "  Cut=" + str(r[cut_col] if cut_col else '?') +
          "  Polish=" + str(r[polish_col] if polish_col else '?') +
          "  Sym=" + str(r[sym_col] if sym_col else '?') +
          "  Fluor=" + str(r[fluor_col] if fluor_col else '?'))
    print("  Location=" + str(r[loc_col] if loc_col else '?') +
          "  Lab=" + str(r[lab_col] if lab_col else '?') +
          "  AgingDays=" + str(r[aging_col] if aging_col else '?') +
          "  Status=" + str(r[status_col] if status_col else '?'))
    print("  Mail Back=" + str(r[mail_back_col] if mail_back_col else '?') +
          "  Final disc=" + str(r[final_disc_col] if final_disc_col else '?') +
          "  Rapnet disc+=" + str(r[rapnet_disc_col] if rapnet_disc_col else '?') +
          "  REAL RAPNET=" + str(r[real_rapnet_col] if real_rapnet_col else '?'))
    print("  Tariff Rap+=" + str(r[tariff_rapnet_col] if tariff_rapnet_col else '?') +
          "  Base pd disc=" + str(r[base_pd_disc_col] if base_pd_disc_col else '?') +
          "  LIMIT=" + str(r[limit_col] if limit_col else '?') +
          "  LIMIT REMARK=" + str(r[limit_remark_col] if limit_remark_col else '?'))
    print("  RapPos IND=" + str(r[rapnet_pos_ind_col] if rapnet_pos_ind_col else '?') +
          "  RapPos World=" + str(r[rapnet_pos_world_col] if rapnet_pos_world_col else '?') +
          "  RapPos USA=" + str(r[rapnet_pos_usa_col] if rapnet_pos_usa_col else '?'))
    if first_comp_col:
        print("  1st comp=" + str(r[first_comp_col]) +
              "  2nd comp=" + str(r[first_comp_col+2] if first_comp_col+2 < len(r) else '?') +
              "  3rd comp=" + str(r[first_comp_col+4] if first_comp_col+4 < len(r) else '?') +
              "  AvgDisc=" + str(r[avg_disc_col] if avg_disc_col else '?'))
    if rapnet_aging_col:
        print("  RAPNET AGING=" + str(r[rapnet_aging_col]) +
              "  Days on Rapnet=" + str(r[days_rapnet_col] if days_rapnet_col else '?'))

# Location breakdown for Round
loc_counts = {}
for r in round_rows:
    loc = r[loc_col] if loc_col else 'Unknown'
    loc_counts[str(loc)] = loc_counts.get(str(loc), 0) + 1
print("\nROUND 0.30-2.00ct by Location:")
for loc, cnt in sorted(loc_counts.items(), key=lambda x: -x[1]):
    print("  " + loc + ": " + str(cnt))

# Size breakdown
size_counts = {}
for r in round_rows:
    s = r[rapsize_col] if rapsize_col else r[cts_col]
    size_counts[str(s)] = size_counts.get(str(s), 0) + 1
print("\nROUND 0.30-2.00ct by RapNet Size bucket:")
for s, cnt in sorted(size_counts.items()):
    print("  " + s + ": " + str(cnt))

wb.close()
print("\nDone.")
