import openpyxl

path = r"E:\Pricing\Monthly\MAY-2026\0.30UP POSITION REPORT - 15-05-2026.xlsx"
print("Checking Polish & Symmetry for ROUND EXCL 0.30-2.00ct stones...")
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb.active

headers = []
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    headers = list(row)

def ci(name):
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None

shape_i  = ci('Shape')
cts_i    = ci('cts')
cut_i    = ci('Cut')
polish_i = ci('Polish')
sym_i    = ci('Symmetry')
color_i  = ci('Color')
clarity_i= ci('Clarity')
fluor_i  = ci('Fluorescence')

# Count Polish/Symmetry combos for ROUND EXCL 0.30-2.00ct
combo_counts = {}
polish_counts = {}
sym_counts = {}
total = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[shape_i] != 'ROUND':
        continue
    try:
        cts = float(row[cts_i])
    except:
        continue
    if not (0.30 <= cts <= 2.00):
        continue
    if str(row[cut_i]).strip() != 'EXCL':
        continue
    total += 1
    polish = str(row[polish_i]).strip() if row[polish_i] else 'None'
    sym    = str(row[sym_i]).strip()    if row[sym_i]    else 'None'
    combo  = polish + ' / ' + sym
    combo_counts[combo]   = combo_counts.get(combo, 0) + 1
    polish_counts[polish] = polish_counts.get(polish, 0) + 1
    sym_counts[sym]       = sym_counts.get(sym, 0) + 1

wb.close()

print("Total ROUND EXCL 0.30-2.00ct stones: " + str(total))
print()
print("Polish distribution:")
for k, v in sorted(polish_counts.items(), key=lambda x: -x[1]):
    pct = round(v/total*100, 1)
    print("  " + k + ": " + str(v) + " (" + str(pct) + "%)")

print()
print("Symmetry distribution:")
for k, v in sorted(sym_counts.items(), key=lambda x: -x[1]):
    pct = round(v/total*100, 1)
    print("  " + k + ": " + str(v) + " (" + str(pct) + "%)")

print()
print("Polish / Symmetry combinations:")
for k, v in sorted(combo_counts.items(), key=lambda x: -x[1]):
    pct = round(v/total*100, 1)
    print("  " + k + ": " + str(v) + " (" + str(pct) + "%)")
