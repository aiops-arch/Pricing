import openpyxl

def quick_peek(path, label, max_rows=8):
    print("\n" + "-"*80)
    print("FILE: " + label)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        print("Sheets: " + str(wb.sheetnames))
        rows_read = 0
        for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
            print("  row" + str(i+1).zfill(2) + ": " + str(list(row)[:20]))
            rows_read += 1
        wb.close()
    except Exception as e:
        print("ERROR: " + str(e))

quick_peek(r"E:\Pricing\Monthly\MAY-2026\0.30UP POSITION REPORT - 15-05-2026.xlsx", "POSITION REPORT 15-05-2026")
quick_peek(r"E:\Pricing\Monthly\JAN-2026\0.30UP POSITION REPORT 01-01-2026.xlsx", "POSITION REPORT 01-01-2026")
