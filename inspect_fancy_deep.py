import openpyxl

path = r"E:\Pricing\FANCY BASE REPORT 23-03-2026.xlsx"
print("Opening Fancy Base Report...")
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb.active

# Get all column headers from row 2
headers = []
for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
    headers = list(row)

print("TOTAL COLUMNS: " + str(len(headers)))
print("\nALL COLUMN HEADERS (showing non-None):")
for i, h in enumerate(headers):
    print("  col" + str(i+1).zfill(3) + ": " + str(h))

# Find ROUND rows and show a sample
print("\n\nSEARCHING FOR ROUND ROWS (showing first 5)...")
round_count = 0
round_samples = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] == 'ROUND':
        round_count += 1
        if round_count <= 5:
            round_samples.append(list(row))

print("Total ROUND rows found: " + str(round_count))
print("\nFirst 5 ROUND rows (first 50 cols):")
for r in round_samples:
    print("  " + str(r[:50]))

# Count all unique shapes
print("\n\nSHAPE COUNTS:")
shape_counts = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    shape = row[0]
    if shape:
        shape_counts[shape] = shape_counts.get(shape, 0) + 1

wb.close()
for shape, count in sorted(shape_counts.items()):
    print("  " + str(shape) + ": " + str(count) + " rows")
